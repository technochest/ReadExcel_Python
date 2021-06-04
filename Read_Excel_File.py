import openpyxl

wb = openpyxl.load_workbook("Enter your Directory Here", data_only=True) #wb represents the excel file

#Overview: This example shows that we are reading some list from an excel document...
#Excel File could contain student names city names, dates etc...
#length of the column could any number of rows....
#Objective: This method should read the excel file and return the names...

def Get_Names(Sheet_Name: str ):
	sht = wb[Sheet_Name]
	_List = []  #To store the names in a list.
	sht_rows = sht.max_row  #toal rows eqalt total rows in the excel sheet
	sht_columns = sht.max_column
	for r in range(2,sht_rows+1):  # 2 is asuming that the column has a header and second row has the actual names.
		for c in range(1,2):
			_List.append(sht.cell(r,c).value)

		return _List